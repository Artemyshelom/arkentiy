"""
Трекер онлайн-оплат: реестр ТБанк vs iiko.

Stateful сверка: каждый онлайн-заказ из iiko отслеживается до подтверждения
реестром ТБанк. Если оплата не подтверждена > N дней — алерт.

Вход: xlsx-файл ежедневного отчёта ТБанк (лист = точка, строки = транзакции).
"""

import html
import json
import logging
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

SECRETS_DIR = Path(__file__).resolve().parent.parent.parent / "secrets"
TBANK_BRANCHES_PATH = SECRETS_DIR / "tbank_branches.json"

OVERDUE_DAYS = 4
TRACKING_START_DATE = "2026-02-18"  # Точка отсчёта: раньше этой даты не смотрим


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class TBankTransaction:
    order_number: str
    amount: float
    commission: float
    order_date: str          # DD.MM.YYYY
    transaction_date: str    # DD.MM.YYYY
    transaction_time: str    # HH:MM:SS
    transaction_id: str
    status: str
    payment_type: str
    delivery_type: str
    payment_order: str       # платёжное поручение


@dataclass
class TBankSheet:
    branch_name: str         # имя листа (Новосибирск_2_Объ)
    ip_name: str
    report_date: str
    tariff: str
    total_commission: float
    total_amount: float
    transactions: list[TBankTransaction] = field(default_factory=list)


@dataclass
class ReconciliationResult:
    confirmed: int = 0
    mismatched: int = 0
    new_pending: int = 0
    missing_in_iiko: int = 0
    total_tbank_orders: int = 0
    total_tbank_amount: float = 0.0
    total_tbank_commission: float = 0.0
    details: list[dict] = field(default_factory=list)
    branch_results: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Auto-detect
# ---------------------------------------------------------------------------

def is_tbank_registry(data: bytes) -> bool:
    """Проверяет, является ли xlsx реестром ТБанк (по структуре шапки)."""
    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                for cell in row:
                    if cell and "Уникальный идентификатор транзакции" in str(cell):
                        wb.close()
                        return True
        wb.close()
    except Exception:
        pass
    return False


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def _parse_date(val) -> str:
    """Преобразует значение ячейки в строку DD.MM.YYYY."""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, str):
        return val.strip()
    return str(val) if val else ""


def _date_to_iso(dd_mm_yyyy: str) -> str:
    """DD.MM.YYYY -> YYYY-MM-DD."""
    try:
        parts = dd_mm_yyyy.split(".")
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except (IndexError, ValueError):
        return dd_mm_yyyy


def parse_tbank_registry(data: bytes) -> list[TBankSheet]:
    """Парсит xlsx реестра ТБанк, возвращает данные по каждому листу (точке)."""
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheets: list[TBankSheet] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows_data) < 7:
            continue

        ip_name = str(rows_data[0][0] or "").strip()
        branch_name = str(rows_data[1][0] or "").strip()
        report_date_raw = str(rows_data[2][0] or "").strip()
        report_date = report_date_raw.replace("Отчетный период ", "").replace("Отчётный период ", "").strip()
        tariff = str(rows_data[3][0] or "").strip()

        total_commission = 0.0
        total_amount = 0.0
        if rows_data[2] and len(rows_data[2]) >= 7:
            total_commission = abs(float(rows_data[2][6] or 0))
        if rows_data[3] and len(rows_data[3]) >= 7:
            total_amount = float(rows_data[3][6] or 0)

        header_row_idx = None
        for i in range(4, min(8, len(rows_data))):
            row = rows_data[i]
            if row and row[0] and "Дата" in str(row[0]) and "приёма" in str(row[0]):
                header_row_idx = i
                break
            if row and row[0] and "Дата приёма" in str(row[0]):
                header_row_idx = i
                break

        if header_row_idx is None:
            logger.warning(f"[tbank] лист '{sheet_name}': не найден заголовок таблицы")
            continue

        data_start = header_row_idx + 1
        if data_start < len(rows_data):
            next_row = rows_data[data_start]
            if next_row and next_row[0] and isinstance(next_row[0], str) and "Приход" in str(next_row[0]):
                data_start += 1

        transactions: list[TBankTransaction] = []
        pending_tx: dict | None = None

        for i in range(data_start, len(rows_data)):
            row = rows_data[i]
            if not row or row[0] is None:
                continue

            first_cell = str(row[0]).strip()
            if first_cell.startswith("Итого") or first_cell.startswith("Оплата на сайте"):
                break

            comment = str(row[9] or "").strip() if len(row) > 9 else ""

            if comment == "Оплата по заказу":
                if pending_tx:
                    transactions.append(TBankTransaction(**pending_tx, commission=0.0))
                tx_status = str(row[5] or "").strip()
                if tx_status != "Закрыт":
                    # Пропускаем незавершённые ("Готовится" и т.п.) — платёж не подтверждён ТБанком
                    pending_tx = None
                    continue
                pending_tx = {
                    "order_date": _parse_date(row[0]),
                    "transaction_date": _parse_date(row[1]),
                    "transaction_id": str(row[2] or ""),
                    "transaction_time": str(row[3] or ""),
                    "order_number": str(int(row[4]) if isinstance(row[4], (int, float)) else row[4] or ""),
                    "status": tx_status,
                    "payment_type": str(row[6] or ""),
                    "delivery_type": str(row[7] or ""),
                    "amount": float(row[8] or 0),
                    "payment_order": str(row[10] or "") if len(row) > 10 else "",
                }
            elif comment == "Комиссия за заказ" and pending_tx:
                commission = abs(float(row[8] or 0))
                transactions.append(TBankTransaction(**pending_tx, commission=commission))
                pending_tx = None

        if pending_tx:
            transactions.append(TBankTransaction(**pending_tx, commission=0.0))

        sheets.append(TBankSheet(
            branch_name=branch_name or sheet_name,
            ip_name=ip_name,
            report_date=report_date,
            tariff=tariff,
            total_commission=total_commission,
            total_amount=total_amount,
            transactions=transactions,
        ))

    wb.close()
    return sheets


# ---------------------------------------------------------------------------
# Branch mapping
# ---------------------------------------------------------------------------

def load_branch_mapping() -> dict[str, str]:
    """Загружает маппинг имён листов ТБанк -> iiko Department."""
    if not TBANK_BRANCHES_PATH.exists():
        return {}
    return json.loads(TBANK_BRANCHES_PATH.read_text(encoding="utf-8"))


def _resolve_branch(sheet_branch: str, mapping: dict[str, str]) -> str | None:
    """Находит iiko Department для имени ветки из реестра ТБанк."""
    if sheet_branch in mapping:
        return mapping[sheet_branch]
    normalized = sheet_branch.replace(" ", "_")
    if normalized in mapping:
        return mapping[normalized]
    for key, value in mapping.items():
        if key.replace("_", " ") == sheet_branch or value == sheet_branch:
            return value
    return None


# ---------------------------------------------------------------------------
# Reconciliation engine
# ---------------------------------------------------------------------------

async def process_registry(
    data: bytes,
    user_id: int = 0,
    chat_id: int = 0,
    filename: str = "",
) -> dict:
    """
    Основная функция обработки реестра ТБанк.
    1. Парсит xlsx
    2. Запрашивает iiko OLAP v2
    3. Регистрирует/подтверждает оплаты в трекере
    4. Формирует отчёт
    """
    from app.clients.iiko_bo_olap_v2 import get_online_orders
    from app.database import (
        confirm_online_payment,
        get_overdue_payments,
        get_pending_payments,
        get_tracking_summary,
        save_tbank_registry_log,
        upsert_online_payment,
    )

    sheets = parse_tbank_registry(data)
    if not sheets:
        return {"error": "Не удалось распарсить файл. Проверьте формат."}

    mapping = load_branch_mapping()
    result = ReconciliationResult()

    all_order_dates: list[str] = []
    for sheet in sheets:
        for tx in sheet.transactions:
            iso = _date_to_iso(tx.order_date)
            if iso not in all_order_dates:
                all_order_dates.append(iso)

    if not all_order_dates:
        return {"error": "В файле нет транзакций."}

    all_order_dates.sort()
    date_from = all_order_dates[0]
    date_to_raw = all_order_dates[-1]
    # +2 дня: iiko может записывать заказы на следующую дату (timezone/midnight cutoff)
    date_to_dt = datetime.fromisoformat(date_to_raw) + timedelta(days=2)
    date_from_extended = max(
        (datetime.fromisoformat(date_from) - timedelta(days=7)).strftime("%Y-%m-%d"),
        TRACKING_START_DATE,
    )

    iiko_orders: dict[str, dict[str, dict]] = {}
    try:
        iiko_orders = await get_online_orders(date_from_extended, date_to_dt.strftime("%Y-%m-%d"))
        total_iiko = sum(len(v) for v in iiko_orders.values())
        logger.info(f"[tbank] iiko онлайн-заказов: {total_iiko} по {len(iiko_orders)} точкам")
    except Exception as e:
        logger.error(f"[tbank] iiko OLAP error: {e}", exc_info=True)

    for dept, orders in iiko_orders.items():
        for order_num, info in orders.items():
            await upsert_online_payment(
                branch=dept,
                order_number=order_num,
                order_date=info.get("date") or date_from,
                iiko_amount=info["amount"],
            )

    report_date = sheets[0].report_date if sheets else ""

    for sheet in sheets:
        iiko_dept = _resolve_branch(sheet.branch_name, mapping)
        if not iiko_dept:
            iiko_dept = sheet.branch_name
            logger.warning(f"[tbank] нет маппинга для '{sheet.branch_name}', используем как есть")

        dept_iiko = iiko_orders.get(iiko_dept, {})
        branch_conf = 0
        branch_mismatch = 0
        branch_missing = 0
        branch_details: list[dict] = []

        for tx in sheet.transactions:
            result.total_tbank_orders += 1
            result.total_tbank_amount += tx.amount
            result.total_tbank_commission += tx.commission

            # Settlement lag: заказы до начала трекинга — всегда будет хвост в начале
            # (платежи до 17:50 МСК → следующий рабочий день, поэтому реестр 18.02
            # содержит заказы от 17.02 и ранее). Сверку для них не проводим.
            if _date_to_iso(tx.order_date) < TRACKING_START_DATE:
                branch_conf += 1
                result.confirmed += 1
                continue

            iiko_entry = dept_iiko.get(tx.order_number)
            iiko_amount = iiko_entry["amount"] if iiko_entry else None

            status = await confirm_online_payment(
                branch=iiko_dept,
                order_number=tx.order_number,
                tbank_amount=tx.amount,
                tbank_commission=tx.commission,
                tbank_confirmed_date=_date_to_iso(tx.transaction_date),
                tbank_transaction_id=tx.transaction_id,
                iiko_amount=iiko_amount,
                tenant_id=1,
            )

            if status in ("confirmed", "created_confirmed"):
                branch_conf += 1
                result.confirmed += 1
            elif status in ("mismatch", "created_mismatch"):
                branch_mismatch += 1
                result.mismatched += 1
                branch_details.append({
                    "type": "mismatch",
                    "order": tx.order_number,
                    "tbank": tx.amount,
                    "iiko": iiko_amount,
                })
            elif "missing_in_iiko" in status:
                branch_missing += 1
                result.missing_in_iiko += 1
                branch_details.append({
                    "type": "missing_in_iiko",
                    "order": tx.order_number,
                    "tbank": tx.amount,
                })

        result.branch_results[iiko_dept] = {
            "sheet_name": sheet.branch_name,
            "confirmed": branch_conf,
            "mismatched": branch_mismatch,
            "missing_in_iiko": branch_missing,
            "total_orders": len(sheet.transactions),
            "total_amount": sum(t.amount for t in sheet.transactions),
            "total_commission": sum(t.commission for t in sheet.transactions),
            "details": branch_details,
        }

    all_pending = await get_pending_payments(since_date=TRACKING_START_DATE, tenant_id=1)
    result.new_pending = len(all_pending)

    overdue = await get_overdue_payments(OVERDUE_DAYS, since_date=TRACKING_START_DATE, tenant_id=1)
    tracking = await get_tracking_summary(since_date=TRACKING_START_DATE, tenant_id=1)

    try:
        await save_tbank_registry_log(
            user_id=user_id,
            chat_id=chat_id,
            filename=filename,
            report_date=report_date,
            total_orders=result.total_tbank_orders,
            confirmed=result.confirmed,
            mismatched=result.mismatched,
            new_pending=result.new_pending,
            missing_in_iiko=result.missing_in_iiko,
        )
    except Exception as e:
        logger.warning(f"[tbank] db log: {e}")

    today_iso = datetime.now(timezone.utc).date().isoformat()
    has_pending = any(p["order_date"] != today_iso for p in all_pending)

    report_text = _build_report(result, overdue, all_pending, tracking, report_date)
    return {
        "report": report_text,
        "result": result,
        "overdue": overdue,
        "all_pending": all_pending,
        "has_pending": has_pending,
    }


# ---------------------------------------------------------------------------
# iiko OLAP query (will be added to iiko_bo_olap_v2.py)
# ---------------------------------------------------------------------------

# The function get_online_orders() is defined in iiko_bo_olap_v2.py


# ---------------------------------------------------------------------------
# Telegram report
# ---------------------------------------------------------------------------

def _fmt_money(val: float) -> str:
    return f"{int(val):,}".replace(",", " ")


def _fmt_date(iso: str) -> str:
    try:
        return datetime.fromisoformat(iso).strftime("%d.%m")
    except Exception:
        return iso


def _days_ago(iso: str) -> int:
    try:
        return (datetime.now(timezone.utc).date() - datetime.fromisoformat(iso).date()).days
    except Exception:
        return 0


def _plural_orders(n: int) -> str:
    if n % 10 == 1 and n % 100 != 11:
        return f"{n} заказ"
    if 2 <= n % 10 <= 4 and not (12 <= n % 100 <= 14):
        return f"{n} заказа"
    return f"{n} заказов"


def _build_report(
    result: ReconciliationResult,
    overdue: list[dict],
    all_pending: list[dict],
    tracking: dict,
    report_date: str,
) -> str:
    lines: list[str] = []
    today_iso = datetime.now(timezone.utc).date().isoformat()

    # ── Блок 1: Текущий реестр ──────────────────────────────────────────────
    lines.append(f"💳 <b>Реестр ТБанк · {html.escape(report_date)}</b>")
    lines.append(f"{result.total_tbank_orders} заказов · {_fmt_money(result.total_tbank_amount)} р · комиссия {_fmt_money(result.total_tbank_commission)} р")

    issues = result.mismatched + result.missing_in_iiko
    if issues == 0:
        lines.append("✅ Все подтверждены")
    else:
        lines.append(f"⚠️ {issues} расхождени{'е' if issues == 1 else 'я' if 2 <= issues <= 4 else 'й'}:")
        for dept, br in result.branch_results.items():
            for d in br.get("details", []):
                if d["type"] == "mismatch":
                    iiko_str = _fmt_money(d["iiko"]) if d["iiko"] is not None else "—"
                    lines.append(f"  #{d['order']} {html.escape(dept)} · ТБанк {_fmt_money(d['tbank'])} р ≠ iiko {iiko_str} р")
                elif d["type"] == "missing_in_iiko":
                    lines.append(f"  #{d['order']} {html.escape(dept)} · {_fmt_money(d['tbank'])} р — не найден в iiko")

    # ── Блок 2: Статус онлайн-оплат ─────────────────────────────────────────
    if tracking:
        lines.append("")
        start_display = _fmt_date(TRACKING_START_DATE)
        lines.append(f"📋 <b>Онлайн-оплаты в iiko · с {start_display}</b>")

        # Pending без сегодня (реестра за сегодня нет по определению)
        non_today_pending = [p for p in all_pending if p["order_date"] != today_iso]
        total_pending_amount = sum(p.get("iiko_amount") or 0 for p in non_today_pending)

        if not non_today_pending:
            lines.append("✅ Все оплаты подтверждены")
        else:
            lines.append(f"Деньги ещё не в банке: {_plural_orders(len(non_today_pending))} на {_fmt_money(total_pending_amount)} р")

            # --- 🔴 Просрочено (> OVERDUE_DAYS) ---
            overdue_groups: dict[tuple, list] = defaultdict(list)
            for p in overdue:
                if p["order_date"] != today_iso:
                    overdue_groups[(p["branch"], p["order_date"])].append(p)

            if overdue_groups:
                lines.append("")
                lines.append("🔴 <b>Просрочено (> 4 дн.)</b>")
                for (branch, date_iso) in sorted(overdue_groups.keys(), key=lambda x: (x[1], x[0])):
                    group = overdue_groups[(branch, date_iso)]
                    amt = sum(p.get("iiko_amount") or 0 for p in group)
                    lines.append(f"{html.escape(branch)} · {_fmt_date(date_iso)} — {_plural_orders(len(group))} ({_fmt_money(amt)} р)")
                    for p in group[:3]:
                        lines.append(f"  └ #{p['order_number']}: {_fmt_money(p.get('iiko_amount') or 0)} р")
                    if len(group) > 3:
                        lines.append(f"  └ ... ещё {len(group) - 3}")

            # --- ⏳ Деньги ещё не в банке (свежие, ≤ OVERDUE_DAYS) ---
            non_overdue_pending = [p for p in non_today_pending if _days_ago(p["order_date"]) < OVERDUE_DAYS]

            pending_groups: dict[tuple, list] = defaultdict(list)
            for p in non_overdue_pending:
                pending_groups[(p["branch"], p["order_date"])].append(p)

            if pending_groups:
                lines.append("")
                lines.append("⏳ <b>Деньги ещё не в банке</b>")
                # Сортировка: больше заказов — выше, потом по branch
                for (branch, date_iso) in sorted(
                    pending_groups.keys(),
                    key=lambda x: (-len(pending_groups[x]), x[0]),
                ):
                    group = pending_groups[(branch, date_iso)]
                    amt = sum(p.get("iiko_amount") or 0 for p in group)
                    lines.append(f"{html.escape(branch)} · {_fmt_date(date_iso)} — {_plural_orders(len(group))} ({_fmt_money(amt)} р)")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Drill-down: branch list (State 1) and branch detail (State 2)
# ---------------------------------------------------------------------------

def build_branch_list(all_pending: list[dict], overdue: list[dict]) -> tuple[str, list]:
    """Возвращает (text, keyboard) для экрана выбора точки (State 1)."""
    today_iso = datetime.now(timezone.utc).date().isoformat()

    branch_overdue: dict[str, list] = defaultdict(list)
    for p in overdue:
        if p["order_date"] != today_iso:
            branch_overdue[p["branch"]].append(p)

    branch_pending: dict[str, list] = defaultdict(list)
    for p in all_pending:
        if p["order_date"] != today_iso and _days_ago(p["order_date"]) < OVERDUE_DAYS:
            branch_pending[p["branch"]].append(p)

    all_branches = set(list(branch_overdue.keys()) + list(branch_pending.keys()))
    back_btn = [{"text": "← Назад к отчёту", "callback_data": "tbank:back"}]

    if not all_branches:
        return "✅ Нет ожидающих заказов", [back_btn]

    def _sort_key(b: str) -> tuple:
        return (-len(branch_overdue.get(b, [])), -len(branch_pending.get(b, [])), b)

    keyboard: list[list[dict]] = []
    for branch in sorted(all_branches, key=_sort_key):
        ov = len(branch_overdue.get(branch, []))
        pend = len(branch_pending.get(branch, []))
        icons = ""
        if ov:
            icons += f"🔴 {ov}"
        if pend:
            icons += f" · ⏳ {pend}" if icons else f"⏳ {pend}"
        label = f"{branch}  {icons}".strip()
        cb = f"tbank:branch:{branch}"
        keyboard.append([{"text": label, "callback_data": cb[:64]}])

    keyboard.append(back_btn)
    text = "📋 <b>Онлайн-оплаты · детализация</b>\n\nВыбери точку:"
    return text, keyboard


def build_branch_detail(branch: str, all_pending: list[dict], overdue: list[dict]) -> tuple[str, list]:
    """Возвращает (text, keyboard) для детального экрана точки (State 2)."""
    today_iso = datetime.now(timezone.utc).date().isoformat()

    br_overdue = [p for p in overdue if p["branch"] == branch and p["order_date"] != today_iso]
    br_pending = [
        p for p in all_pending
        if p["branch"] == branch and p["order_date"] != today_iso and _days_ago(p["order_date"]) < OVERDUE_DAYS
    ]

    keyboard = [[
        {"text": "← К точкам", "callback_data": "tbank:branches"},
        {"text": "← К отчёту", "callback_data": "tbank:back"},
    ]]

    if not br_overdue and not br_pending:
        return f"📋 <b>{html.escape(branch)}</b>\n\n✅ Нет ожидающих заказов", keyboard

    lines = [f"📋 <b>{html.escape(branch)}</b> · ожидают подтверждения"]

    def _render_group(items: list[dict]) -> None:
        by_date: dict[str, list] = defaultdict(list)
        for p in items:
            by_date[p["order_date"]].append(p)
        for date_iso in sorted(by_date.keys()):
            group = by_date[date_iso]
            amt = sum(p.get("iiko_amount") or 0 for p in group)
            lines.append(f"{_fmt_date(date_iso)} — {_plural_orders(len(group))} · {_fmt_money(amt)} р")
            for p in group:
                lines.append(f"  └ #{p['order_number']}: {_fmt_money(p.get('iiko_amount') or 0)} р")

    if br_overdue:
        lines.append("")
        lines.append("🔴 <b>Просрочено (деньги не пришли):</b>")
        _render_group(br_overdue)

    if br_pending:
        lines.append("")
        lines.append("⏳ <b>Деньги ещё не в банке:</b>")
        _render_group(br_pending)

    return "\n".join(lines), keyboard


# ---------------------------------------------------------------------------
# Payout report (Ежедневный отчёт по выплатам)
# ---------------------------------------------------------------------------

PAYOUT_DELAY_DAYS = 2  # дней после tbank_confirmed_date до алерта о задержке выплаты


@dataclass
class TBankPayoutTransaction:
    order_number: str
    payment_date: str     # DD.MM.YYYY
    payout_date: str      # DD.MM.YYYY
    amount: float         # Сумма операции
    commission: float     # Комиссия МП
    net_amount: float     # К перечислению
    operation_type: str   # Debit / Credit
    payment_system: str   # SBP / Mir / Visa / Mastercard


@dataclass
class TBankPayoutSheet:
    branch_name: str
    payout_date: str
    total_amount: float
    total_commission: float
    total_net: float
    transactions: list[TBankPayoutTransaction] = field(default_factory=list)


def is_tbank_payout(data: bytes) -> bool:
    """Проверяет, является ли xlsx отчётом по выплатам ТБанк."""
    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                for cell in row:
                    if cell and "Выплата от" in str(cell):
                        wb.close()
                        return True
        wb.close()
    except Exception:
        pass
    return False


def parse_tbank_payout(data: bytes) -> list[TBankPayoutSheet]:
    """Парсит xlsx отчёта по выплатам ТБанк."""
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheets: list[TBankPayoutSheet] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
        if len(rows_data) < 8:
            continue

        # Извлекаем дату выплаты из строки 1: "... Выплата от DD.MM.YYYY."
        header_text = str(rows_data[0][0] or "")
        payout_date_str = ""
        import re as _re
        m = _re.search(r"Выплата от (\d{2}\.\d{2}\.\d{4})", header_text)
        if m:
            payout_date_str = m.group(1)

        branch_name = str(rows_data[1][0] or "").strip()
        total_amount = float(rows_data[2][1] or 0)
        total_commission = abs(float(rows_data[3][1] or 0))
        total_net = float(rows_data[4][1] or 0)

        if total_amount == 0 and total_net == 0:
            # Пустой лист — пропускаем
            continue

        transactions: list[TBankPayoutTransaction] = []
        # Строка 8 (индекс 7) — заголовок, данные с строки 9 (индекс 8)
        for row in rows_data[8:]:
            if not row or row[0] is None:
                continue
            try:
                order_number = str(int(row[3]) if isinstance(row[3], (int, float)) else row[3] or "")
                if not order_number:
                    continue
                transactions.append(TBankPayoutTransaction(
                    order_number=order_number,
                    payment_date=_parse_date(row[4]),
                    payout_date=_parse_date(row[9]) if row[9] else payout_date_str,
                    amount=float(row[11] or 0),
                    commission=abs(float(row[12] or 0)),
                    net_amount=float(row[13] or 0),
                    operation_type=str(row[14] or "Debit").strip(),
                    payment_system=str(row[2] or "").strip(),
                ))
            except (IndexError, ValueError, TypeError):
                continue

        sheets.append(TBankPayoutSheet(
            branch_name=branch_name or sheet_name,
            payout_date=payout_date_str,
            total_amount=total_amount,
            total_commission=total_commission,
            total_net=total_net,
            transactions=transactions,
        ))

    wb.close()
    return sheets


async def process_payout(
    data: bytes,
    user_id: int = 0,
    chat_id: int = 0,
    filename: str = "",
) -> dict:
    """Обработка отчёта по выплатам ТБанк."""
    from app.database import confirm_payout, get_payout_delayed, record_chargeback

    sheets = parse_tbank_payout(data)
    if not sheets:
        return {"error": "В файле нет данных о выплатах."}

    mapping = load_branch_mapping()
    payout_date = sheets[0].payout_date if sheets else ""

    confirmed_count = 0
    chargeback_list: list[dict] = []
    not_found_list: list[dict] = []
    total_amount = 0.0
    total_net = 0.0

    for sheet in sheets:
        iiko_dept = _resolve_branch(sheet.branch_name, mapping) or sheet.branch_name
        total_amount += sheet.total_amount
        total_net += sheet.total_net

        for tx in sheet.transactions:
            tx_payout_iso = _date_to_iso(tx.payout_date) if tx.payout_date else ""

            if tx_payout_iso and tx_payout_iso < TRACKING_START_DATE:
                continue

            if tx.operation_type.lower() == "credit":
                chargeback_list.append({
                    "branch": iiko_dept,
                    "order": tx.order_number,
                    "date": tx.payment_date,
                    "amount": tx.amount,
                })
                await record_chargeback(
                    branch=iiko_dept,
                    order_number=tx.order_number,
                    chargeback_date=tx_payout_iso or payout_date,
                    amount=tx.amount,
                )
            else:
                result = await confirm_payout(
                    branch=iiko_dept,
                    order_number=tx.order_number,
                    payout_date=tx_payout_iso or _date_to_iso(payout_date),
                    payout_amount=tx.net_amount,
                )
                if result == "confirmed":
                    confirmed_count += 1
                else:
                    not_found_list.append({
                        "branch": iiko_dept,
                        "order": tx.order_number,
                        "amount": tx.net_amount,
                    })

    delayed = await get_payout_delayed(PAYOUT_DELAY_DAYS, since_date=TRACKING_START_DATE, tenant_id=1)
    report = _build_payout_report(sheets, payout_date, confirmed_count, chargeback_list, delayed, not_found_list, total_amount, total_net)
    return {"report": report}


def _build_payout_report(
    sheets: list[TBankPayoutSheet],
    payout_date: str,
    confirmed: int,
    chargebacks: list[dict],
    delayed: list[dict],
    not_found: list[dict],
    total_amount: float,
    total_net: float,
) -> str:
    lines: list[str] = []

    lines.append(f"💸 <b>Выплата ТБанк · {html.escape(payout_date)}</b>")
    lines.append(f"{_fmt_money(total_amount)} р · на счёт: {_fmt_money(total_net)} р")

    # По точкам
    for sheet in sheets:
        if sheet.total_amount > 0:
            lines.append(f"  {html.escape(sheet.branch_name)} — {_fmt_money(sheet.total_net)} р")

    lines.append("")
    if confirmed:
        lines.append(f"✅ Подтверждено выплат: {_plural_orders(confirmed)}")

    # Возвраты/чарджбэки
    if chargebacks:
        lines.append("")
        lines.append(f"💥 <b>Возвраты ({len(chargebacks)}):</b>")
        for cb in chargebacks:
            lines.append(f"  #{cb['order']} {html.escape(cb['branch'])} · {_fmt_date(_date_to_iso(cb['date']))} — -{_fmt_money(cb['amount'])} р")

    # Задержка перечисления
    if delayed:
        lines.append("")
        lines.append(f"⏰ <b>Задержка перечисления (> {PAYOUT_DELAY_DAYS} дн.):</b>")
        delayed_groups: dict[tuple, list] = defaultdict(list)
        for p in delayed:
            delayed_groups[(p["branch"], p.get("tbank_confirmed_date", "")[:10])].append(p)
        for (branch, conf_date) in sorted(delayed_groups.keys(), key=lambda x: x[1]):
            group = delayed_groups[(branch, conf_date)]
            amt = sum(p.get("tbank_amount") or p.get("iiko_amount") or 0 for p in group)
            lines.append(f"  {html.escape(branch)} · подтв. {_fmt_date(conf_date)} — {_plural_orders(len(group))} ({_fmt_money(amt)} р)")
    elif confirmed:
        lines.append("⏰ Задержек нет")

    return "\n".join(lines)
