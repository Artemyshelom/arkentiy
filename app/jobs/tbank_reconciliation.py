"""
Трекер онлайн-оплат: реестр ТБанк vs iiko.

Stateful сверка: каждый онлайн-заказ из iiko отслеживается до подтверждения
реестром ТБанк. Если оплата не подтверждена > N дней — алерт.

Вход: xlsx-файл ежедневного отчёта ТБанк (лист = точка, строки = транзакции).
"""

import html
import json
import logging
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

SECRETS_DIR = Path(__file__).resolve().parent.parent.parent / "secrets"
TBANK_BRANCHES_PATH = SECRETS_DIR / "tbank_branches.json"

OVERDUE_DAYS = 4


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class TBankTransaction:
    order_number: str
    amount: float
    commission: float
    order_date: str          # DD.MM.YYYY
    transaction_date: str    # DD.MM.YYYY
    transaction_time: str    # HH:MM:SS
    transaction_id: str
    status: str
    payment_type: str
    delivery_type: str
    payment_order: str       # платёжное поручение


@dataclass
class TBankSheet:
    branch_name: str         # имя листа (Новосибирск_2_Объ)
    ip_name: str
    report_date: str
    tariff: str
    total_commission: float
    total_amount: float
    transactions: list[TBankTransaction] = field(default_factory=list)


@dataclass
class ReconciliationResult:
    confirmed: int = 0
    mismatched: int = 0
    new_pending: int = 0
    missing_in_iiko: int = 0
    total_tbank_orders: int = 0
    total_tbank_amount: float = 0.0
    total_tbank_commission: float = 0.0
    details: list[dict] = field(default_factory=list)
    branch_results: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Auto-detect
# ---------------------------------------------------------------------------

def is_tbank_registry(data: bytes) -> bool:
    """Проверяет, является ли xlsx реестром ТБанк (по структуре шапки)."""
    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                for cell in row:
                    if cell and "Уникальный идентификатор транзакции" in str(cell):
                        wb.close()
                        return True
        wb.close()
    except Exception:
        pass
    return False


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def _parse_date(val) -> str:
    """Преобразует значение ячейки в строку DD.MM.YYYY."""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, str):
        return val.strip()
    return str(val) if val else ""


def _date_to_iso(dd_mm_yyyy: str) -> str:
    """DD.MM.YYYY -> YYYY-MM-DD."""
    try:
        parts = dd_mm_yyyy.split(".")
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except (IndexError, ValueError):
        return dd_mm_yyyy


def parse_tbank_registry(data: bytes) -> list[TBankSheet]:
    """Парсит xlsx реестра ТБанк, возвращает данные по каждому листу (точке)."""
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheets: list[TBankSheet] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows_data) < 7:
            continue

        ip_name = str(rows_data[0][0] or "").strip()
        branch_name = str(rows_data[1][0] or "").strip()
        report_date_raw = str(rows_data[2][0] or "").strip()
        report_date = report_date_raw.replace("Отчетный период ", "").replace("Отчётный период ", "").strip()
        tariff = str(rows_data[3][0] or "").strip()

        total_commission = 0.0
        total_amount = 0.0
        if rows_data[2] and len(rows_data[2]) >= 7:
            total_commission = abs(float(rows_data[2][6] or 0))
        if rows_data[3] and len(rows_data[3]) >= 7:
            total_amount = float(rows_data[3][6] or 0)

        header_row_idx = None
        for i in range(4, min(8, len(rows_data))):
            row = rows_data[i]
            if row and row[0] and "Дата" in str(row[0]) and "приёма" in str(row[0]):
                header_row_idx = i
                break
            if row and row[0] and "Дата приёма" in str(row[0]):
                header_row_idx = i
                break

        if header_row_idx is None:
            logger.warning(f"[tbank] лист '{sheet_name}': не найден заголовок таблицы")
            continue

        data_start = header_row_idx + 1
        if data_start < len(rows_data):
            next_row = rows_data[data_start]
            if next_row and next_row[0] and isinstance(next_row[0], str) and "Приход" in str(next_row[0]):
                data_start += 1

        transactions: list[TBankTransaction] = []
        pending_tx: dict | None = None

        for i in range(data_start, len(rows_data)):
            row = rows_data[i]
            if not row or row[0] is None:
                continue

            first_cell = str(row[0]).strip()
            if first_cell.startswith("Итого") or first_cell.startswith("Оплата на сайте"):
                break

            comment = str(row[9] or "").strip() if len(row) > 9 else ""

            if comment == "Оплата по заказу":
                if pending_tx:
                    transactions.append(TBankTransaction(**pending_tx, commission=0.0))
                pending_tx = {
                    "order_date": _parse_date(row[0]),
                    "transaction_date": _parse_date(row[1]),
                    "transaction_id": str(row[2] or ""),
                    "transaction_time": str(row[3] or ""),
                    "order_number": str(int(row[4]) if isinstance(row[4], (int, float)) else row[4] or ""),
                    "status": str(row[5] or ""),
                    "payment_type": str(row[6] or ""),
                    "delivery_type": str(row[7] or ""),
                    "amount": float(row[8] or 0),
                    "payment_order": str(row[10] or "") if len(row) > 10 else "",
                }
            elif comment == "Комиссия за заказ" and pending_tx:
                commission = abs(float(row[8] or 0))
                transactions.append(TBankTransaction(**pending_tx, commission=commission))
                pending_tx = None

        if pending_tx:
            transactions.append(TBankTransaction(**pending_tx, commission=0.0))

        sheets.append(TBankSheet(
            branch_name=branch_name or sheet_name,
            ip_name=ip_name,
            report_date=report_date,
            tariff=tariff,
            total_commission=total_commission,
            total_amount=total_amount,
            transactions=transactions,
        ))

    wb.close()
    return sheets


# ---------------------------------------------------------------------------
# Branch mapping
# ---------------------------------------------------------------------------

def load_branch_mapping() -> dict[str, str]:
    """Загружает маппинг имён листов ТБанк -> iiko Department."""
    if not TBANK_BRANCHES_PATH.exists():
        return {}
    return json.loads(TBANK_BRANCHES_PATH.read_text(encoding="utf-8"))


def _resolve_branch(sheet_branch: str, mapping: dict[str, str]) -> str | None:
    """Находит iiko Department для имени ветки из реестра ТБанк."""
    if sheet_branch in mapping:
        return mapping[sheet_branch]
    normalized = sheet_branch.replace(" ", "_")
    if normalized in mapping:
        return mapping[normalized]
    for key, value in mapping.items():
        if key.replace("_", " ") == sheet_branch or value == sheet_branch:
            return value
    return None


# ---------------------------------------------------------------------------
# Reconciliation engine
# ---------------------------------------------------------------------------

async def process_registry(
    data: bytes,
    user_id: int = 0,
    chat_id: int = 0,
    filename: str = "",
) -> dict:
    """
    Основная функция обработки реестра ТБанк.
    1. Парсит xlsx
    2. Запрашивает iiko OLAP v2
    3. Регистрирует/подтверждает оплаты в трекере
    4. Формирует отчёт
    """
    from app.clients.iiko_bo_olap_v2 import get_online_orders
    from app.database import (
        confirm_online_payment,
        get_overdue_payments,
        get_pending_payments,
        save_tbank_registry_log,
        upsert_online_payment,
    )

    sheets = parse_tbank_registry(data)
    if not sheets:
        return {"error": "Не удалось распарсить файл. Проверьте формат."}

    mapping = load_branch_mapping()
    result = ReconciliationResult()

    all_order_dates: list[str] = []
    for sheet in sheets:
        for tx in sheet.transactions:
            iso = _date_to_iso(tx.order_date)
            if iso not in all_order_dates:
                all_order_dates.append(iso)

    if not all_order_dates:
        return {"error": "В файле нет транзакций."}

    all_order_dates.sort()
    date_from = all_order_dates[0]
    date_to_raw = all_order_dates[-1]
    date_to_dt = datetime.fromisoformat(date_to_raw) + timedelta(days=1)
    date_from_extended = (datetime.fromisoformat(date_from) - timedelta(days=7)).strftime("%Y-%m-%d")

    iiko_orders: dict[str, dict[str, float]] = {}
    try:
        iiko_orders = await get_online_orders(date_from_extended, date_to_dt.strftime("%Y-%m-%d"))
        logger.info(f"[tbank] iiko онлайн-заказов: {sum(len(v) for v in iiko_orders.values())} по {len(iiko_orders)} точкам")
    except Exception as e:
        logger.error(f"[tbank] iiko OLAP error: {e}", exc_info=True)

    for dept, orders in iiko_orders.items():
        for order_num, amount in orders.items():
            await upsert_online_payment(
                branch=dept,
                order_number=order_num,
                order_date=date_from,
                iiko_amount=amount,
            )

    report_date = sheets[0].report_date if sheets else ""

    for sheet in sheets:
        iiko_dept = _resolve_branch(sheet.branch_name, mapping)
        if not iiko_dept:
            iiko_dept = sheet.branch_name
            logger.warning(f"[tbank] нет маппинга для '{sheet.branch_name}', используем как есть")

        dept_iiko = iiko_orders.get(iiko_dept, {})
        branch_conf = 0
        branch_mismatch = 0
        branch_missing = 0
        branch_details: list[dict] = []

        for tx in sheet.transactions:
            result.total_tbank_orders += 1
            result.total_tbank_amount += tx.amount
            result.total_tbank_commission += tx.commission

            iiko_amount = dept_iiko.get(tx.order_number)

            status = await confirm_online_payment(
                branch=iiko_dept,
                order_number=tx.order_number,
                tbank_amount=tx.amount,
                tbank_commission=tx.commission,
                tbank_confirmed_date=_date_to_iso(tx.transaction_date),
                tbank_transaction_id=tx.transaction_id,
                iiko_amount=iiko_amount,
            )

            if status in ("confirmed", "created_confirmed"):
                branch_conf += 1
                result.confirmed += 1
            elif status in ("mismatch", "created_mismatch"):
                branch_mismatch += 1
                result.mismatched += 1
                branch_details.append({
                    "type": "mismatch",
                    "order": tx.order_number,
                    "tbank": tx.amount,
                    "iiko": iiko_amount,
                })
            elif "missing_in_iiko" in status:
                branch_missing += 1
                result.missing_in_iiko += 1
                branch_details.append({
                    "type": "missing_in_iiko",
                    "order": tx.order_number,
                    "tbank": tx.amount,
                })

        result.branch_results[iiko_dept] = {
            "sheet_name": sheet.branch_name,
            "confirmed": branch_conf,
            "mismatched": branch_mismatch,
            "missing_in_iiko": branch_missing,
            "total_orders": len(sheet.transactions),
            "total_amount": sum(t.amount for t in sheet.transactions),
            "total_commission": sum(t.commission for t in sheet.transactions),
            "details": branch_details,
        }

    all_pending = await get_pending_payments()
    result.new_pending = len(all_pending)

    overdue = await get_overdue_payments(OVERDUE_DAYS)

    try:
        await save_tbank_registry_log(
            user_id=user_id,
            chat_id=chat_id,
            filename=filename,
            report_date=report_date,
            total_orders=result.total_tbank_orders,
            confirmed=result.confirmed,
            mismatched=result.mismatched,
            new_pending=result.new_pending,
            missing_in_iiko=result.missing_in_iiko,
        )
    except Exception as e:
        logger.warning(f"[tbank] db log: {e}")

    report_text = _build_report(result, overdue, report_date)
    return {
        "report": report_text,
        "result": result,
        "overdue": overdue,
    }


# ---------------------------------------------------------------------------
# iiko OLAP query (will be added to iiko_bo_olap_v2.py)
# ---------------------------------------------------------------------------

# The function get_online_orders() is defined in iiko_bo_olap_v2.py


# ---------------------------------------------------------------------------
# Telegram report
# ---------------------------------------------------------------------------

def _fmt_money(val: float) -> str:
    return f"{int(val):,}".replace(",", " ")


def _build_report(result: ReconciliationResult, overdue: list[dict], report_date: str) -> str:
    lines: list[str] = []

    lines.append(f"<b>СВЕРКА ОНЛАЙН-ОПЛАТ | реестр {html.escape(report_date)}</b>")
    lines.append("")

    for dept, br in result.branch_results.items():
        ok = br["mismatched"] == 0 and br["missing_in_iiko"] == 0
        icon = "\u2705" if ok else "\u26a0\ufe0f"
        lines.append(f"{icon} <b>{html.escape(dept)}</b>")
        lines.append(f"  \u0422\u0411\u0430\u043d\u043a: {br['total_orders']} \u0437\u0430\u043a., {_fmt_money(br['total_amount'])} \u0440")
        lines.append(f"  \u041a\u043e\u043c\u0438\u0441\u0441\u0438\u044f: {_fmt_money(br['total_commission'])} \u0440")
        lines.append(f"  \u041f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d\u043e: {br['confirmed']}")

        for d in br["details"]:
            if d["type"] == "mismatch":
                iiko_str = _fmt_money(d['iiko']) if d['iiko'] is not None else "\u2014"
                lines.append(f"  \u26a0\ufe0f #{d['order']}: \u0422\u0411\u0430\u043d\u043a {_fmt_money(d['tbank'])} \u0440 vs iiko {iiko_str} \u0440")
            elif d["type"] == "missing_in_iiko":
                lines.append(f"  \u2753 #{d['order']}: \u0422\u0411\u0430\u043d\u043a {_fmt_money(d['tbank'])} \u0440 \u2014 \u043d\u0435\u0442 \u0432 iiko")
        lines.append("")

    lines.append(f"<b>\u0418\u0422\u041e\u0413\u041e:</b> {result.total_tbank_orders} \u0437\u0430\u043a\u0430\u0437\u043e\u0432, {_fmt_money(result.total_tbank_amount)} \u0440")
    lines.append(f"\u041a\u043e\u043c\u0438\u0441\u0441\u0438\u044f \u0422\u0411\u0430\u043d\u043a: {_fmt_money(result.total_tbank_commission)} \u0440")
    issues = result.mismatched + result.missing_in_iiko
    if issues == 0:
        lines.append("\u2705 \u0412\u0441\u0435 \u0437\u0430\u043a\u0430\u0437\u044b \u043f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d\u044b")
    else:
        lines.append(f"\u26a0\ufe0f \u0420\u0430\u0441\u0445\u043e\u0436\u0434\u0435\u043d\u0438\u0439: {issues}")

    if overdue:
        lines.append("")
        lines.append(f"\u203c\ufe0f <b>\u041f\u0420\u041e\u0421\u0420\u041e\u0427\u0415\u041d\u041d\u042b\u0415 \u041e\u041f\u041b\u0410\u0422\u042b (> {OVERDUE_DAYS} \u0434\u043d\u0435\u0439):</b>")
        by_date: dict[str, list] = defaultdict(list)
        for p in overdue:
            by_date[p["order_date"]].append(p)

        for dt in sorted(by_date.keys()):
            payments = by_date[dt]
            total = sum(p["iiko_amount"] for p in payments)
            try:
                display_date = datetime.fromisoformat(dt).strftime("%d.%m")
            except Exception:
                display_date = dt
            days_ago = (datetime.now(timezone.utc).date() - datetime.fromisoformat(dt).date()).days
            lines.append(f"  \U0001f534 {display_date} ({days_ago} \u0434\u043d.): {len(payments)} \u0437\u0430\u043a., {_fmt_money(total)} \u0440")
            for p in payments[:5]:
                lines.append(f"    #{p['order_number']} {html.escape(p['branch'])} \u2014 {_fmt_money(p['iiko_amount'])} \u0440")
            if len(payments) > 5:
                lines.append(f"    ... \u0438 \u0435\u0449\u0451 {len(payments) - 5}")

    pending_all_count = result.new_pending
    if pending_all_count > 0 and not overdue:
        lines.append("")
        lines.append(f"\u23f3 \u041e\u0436\u0438\u0434\u0430\u044e\u0442 \u043f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d\u0438\u044f: {pending_all_count} \u0437\u0430\u043a\u0430\u0437\u043e\u0432")
    elif pending_all_count > 0:
        lines.append(f"\u23f3 \u0412\u0441\u0435\u0433\u043e \u043e\u0436\u0438\u0434\u0430\u044e\u0442: {pending_all_count} \u0437\u0430\u043a\u0430\u0437\u043e\u0432")

    return "\n".join(lines)
